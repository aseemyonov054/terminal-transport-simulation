import simpy
import random
import openpyxl

def calculate_system_util(total_model_time, total_time, gates):
    '''
        calculates systems statistics (a service utillization and waiting time of a car)
    '''
    result_util = [0] * len(gates)
    result_wait = [0] * len(gates)
    for k in range(len(gates)):
        total_cars_in_service = 0
        total_cars_in_wait = 0
        for _, car_stats in total_time.items():
            total_cars_in_service += car_stats[0][k]
            total_cars_in_wait += car_stats[1][k]
        result_util[k] = total_cars_in_service / (total_model_time * gates[k].capacity)
        result_wait[k] = total_cars_in_wait / len(total_time.keys())
    return result_util, result_wait

def calculate_average_queue_length(total_model_time, queue_length, queue_point):
    '''
        Calculates average queue length for each server
    '''
    average_queue_length = [0] * len(queue_point)
    for i in range(1, len(queue_length)):
        for ch in range(len(queue_point)):
            average_queue_length[ch] += queue_length[i-1][1 + ch] * (queue_length[i][0] - queue_length[i-1][0])
    for ch in range(len(queue_point)):
        average_queue_length[ch] /= total_model_time
    return average_queue_length

def car(env, name, i, gates, list_of_events, total_time, queue_length, list_of_points, queue_point, road_travel_time, 
        sheet, travel_time, time_in_gates):
    '''
        Process of car service on terminal
    '''
    # variables for stats
    list_of_events[name] = []
    waiting_time = [0] * len(gates)
    service_time = [0] * len(gates)
    # start process
    arrive = env.now
    point_counter = 0
    list_of_events[name].append(["delay", arrive, list_of_points[0][1]],)
    service_position = 0
    sheet.cell(row = i+2, column=1).value = i
    sheet.cell(row=i+2, column=2).value = arrive
    var_queue_length = []
    for k in range(1, len(list_of_points)):
        if list_of_points[k][0] == "move":
            # перемещение к точке
            yield env.timeout(road_travel_time[k-1])
            to_point1 = env.now
            travle_time1 = to_point1 - arrive
            if len(list_of_points[k-1]) == 3:
                list_of_events[name].append(["rotate", 0.01, list_of_points[k-1][2]],)
            list_of_events[name].append(["move", travle_time1, list_of_points[k][1]],)
            arrive = env.now
        else:
            # перемещение к точке
            yield env.timeout(road_travel_time[k-1])
            gates_business = gates[point_counter].count
            with gates[point_counter].request() as req1:
                to_point1 = env.now
                travle_time1 = to_point1 - arrive
                if len(list_of_points[k]) == 3:
                    list_of_events[name].append(["rotate", 0.01, list_of_points[k][2]],)
                short_queue_lenght = [round(to_point1,1)]
                for ch in range(len(gates)):
                    short_queue_lenght.append(len(gates[ch].queue))
                list_of_events[name].append(["move", travle_time1, list_of_points[k][1], short_queue_lenght],)
                # очередь
                for y in range(len(gates[point_counter].queue_places)):
                    if gates[point_counter].queue_places[y] == 0:
                        queue_position = y
                        var_queue_length = [to_point1]
                        for ch in range(len(gates)):
                            var_queue_length.append(gates[ch].num_in_queue)
                        queue_length.append(var_queue_length)
                        gates[point_counter].queue_places[y] = name
                        break
                for y in range(len(gates[point_counter].busy_service)):
                    if gates[point_counter].busy_service[y] == 0:
                        service_position = y
                        gates[point_counter].busy_service[y] = name
                        break
                yield req1
                #gates[point_counter].release(req1)
                req1.name = name
                from_queue1 = env.now
                if len(gates[point_counter].queue) == 0 and gates_business < gates[point_counter].capacity:
                    queue_position = 0
                else:
                    queue_delay1 = from_queue1 - to_point1
                    waiting_time[point_counter] += queue_delay1
                    list_of_events[name].append(["delay", queue_delay1, [queue_point[point_counter][0][0], queue_point[point_counter][0][1] + 15 * queue_position]],)
                    #list_of_events[name].append(["delay", queue_delay1, [queue_point[k][0][0] - 50 * random.random(), queue_point[k][0][1] + 50 * random.random()]],)
                yield env.timeout(random.expovariate(1 / gates[point_counter].service_time))
                arrive = env.now
                time_in_service1 = arrive - from_queue1
                service_time[point_counter] += time_in_service1
                short_queue_lenght = [round(env.now, 1)]
                for ch in range(len(gates)):
                    short_queue_lenght.append(len(gates[ch].queue))
                if list_of_points[k][2] == 0:
                    list_of_events[name].append(["delay", time_in_service1, [list_of_points[k][1][0] + 8 * service_position, list_of_points[k][1][1]], short_queue_lenght],)
                else:
                    list_of_events[name].append(["delay", time_in_service1, [list_of_points[k][1][0], list_of_points[k][1][1]+ 8 * service_position], short_queue_lenght],)
                for c in range(len(gates[point_counter].busy_service)):
                    if gates[point_counter].busy_service[c] == name:
                        gates[point_counter].busy_service[c] = 0
                        break
                for c in range(len(gates[point_counter].queue_places)):
                    if gates[point_counter].queue_places[c] == name:
                        gates[point_counter].queue_places[c] = 0
                        break
                var_queue_length = [from_queue1]
                for ch in range(len(gates)):
                    var_queue_length.append(len(gates[ch].queue))
                queue_length.append(var_queue_length)
            point_counter += 1
    total_time[name] = [service_time, waiting_time]
    sheet.cell(row=i+2, column=3).value = env.now
    for k1 in range(len(service_time)):
        sheet.cell(row=i+2, column=4+k1).value = service_time[k1]
    for k2 in range(len(waiting_time)):
        sheet.cell(row=i+2, column=7+k2).value = waiting_time[k2]
    for k3 in range(len(var_queue_length)):
        sheet.cell(row=i+2, column=10+k3).value = var_queue_length[k3]

def source(env, num_of_cars, interval, gates, list_of_processes, sheet):
    '''
        creates process of calls' processing
    '''
    #random.seed(42)
    for i in range(num_of_cars):
        for j in range(len(list_of_processes)):
            pr = list_of_processes[j]
            env.process(car(env, "Car_" + str(i), i+2, gates[j], pr.list_of_events, pr.total_time, pr.queue_length, pr.road_path, 
                    pr.queue_points, pr.travel_time, sheet, travel_time=5, time_in_gates=20))
            t = random.expovariate(1 / interval)
            yield env.timeout(t)

def run_simulation(num_of_calls, list_of_processes, filename):
    '''
        function start the simulation
    '''
    # data initiallization
    gates = []
    env = simpy.Environment()
    for process_descr in list_of_processes:
        #list_of_points = process_descr.road_path
        queue_point = process_descr.queue_points
        #road_travel_time = process_descr.travel_time
        # result dict
        #list_of_events = process_descr.list_of_events
        #total_time = {}
        #queue_length = []
        # simulation
        # create gates
        var_gates = []
        for i in range(len(queue_point)):
            var_gates.append(simpy.Resource(env, capacity=queue_point[i][1]))
            var_gates[i].busy_service = [0] * queue_point[i][1]
            var_gates[i].queue_places = [0] * num_of_calls
            var_gates[i].service_time = queue_point[i][2]
            var_gates[i].num_in_queue = 0
        #process_descr.gates = gates
        gates.append(var_gates)
        #run simulation
        #env.process(source(env, num_of_calls, 10, gates, process_descr.list_of_events, total_time, queue_length, list_of_points, queue_point, road_travel_time))
    arrive_interval = 24 * 60 / num_of_calls
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.cell(row=3, column=1).value = "#"
    sheet.cell(row=3, column=2).value = "time_in"
    sheet.cell(row=3, column=3).value = "time_out"
    sheet.cell(row=3, column=4).value = "serve_1"
    sheet.cell(row=3, column=5).value = "serve_2"
    sheet.cell(row=3, column=6).value = "serve_3"
    sheet.cell(row=3, column=7).value = "wait_1"
    sheet.cell(row=3, column=8).value = "wait_2"
    sheet.cell(row=3, column=9).value = "wait_3"
    sheet.cell(row=3, column=10).value = "time"
    sheet.cell(row=3, column=11).value = "queue_1"
    sheet.cell(row=3, column=12).value = "queue_2"
    sheet.cell(row=3, column=13).value = "queue_3"
    env.process(source(env, num_of_calls, arrive_interval, gates, list_of_processes, sheet))
    env.run()
    wb.save(filename)
    # collect statistics
    total_model_time = env.now
    '''gates_utillization, car_waiting_time = calculate_system_util(total_model_time, total_time, gates)
    average_queue_length = calculate_average_queue_length(total_model_time, queue_length, queue_point)
    print(gates_utillization)
    print(car_waiting_time)
    print(queue_length)
    print(average_queue_length)'''
    return list_of_processes